#  
from joblib import dump, load
from app.objects.manager import Lithologie
from codecarbon import EmissionsTracker
from PIL import Image
import os
import pandas as pd
import numpy as np
import glob
import tensorflow as tf
from tensorflow.keras.preprocessing.image import load_img, img_to_array
from tensorflow.keras.applications.resnet50 import preprocess_input
from tensorflow.keras.preprocessing.image import ImageDataGenerator
import openpyxl
from openpyxl.chart import PieChart, Reference
import pickle
import random
from sklearn.model_selection import train_test_split
from tensorflow.keras import Input, Sequential, Model
from tensorflow.keras import backend as K
from tensorflow.keras.callbacks import EarlyStopping
from tensorflow.keras.layers import Conv2D, MaxPooling2D, Flatten, Dense, Lambda, BatchNormalization, Activation, Dropout
from tensorflow.keras.regularizers import l2

class SiameseNetwork(object):
    def __init__(self, seed, width, height, cells, loss, metrics, optimizer, dropout_rate):
        """
        Seed - The seed used to initialize the weights
        width, height, cells - used for defining the tensors used for the input images
        loss, metrics, optimizer, dropout_rate - settings used for compiling the siamese model (e.g., 'Accuracy' and 'ADAM)
        """
        K.clear_session()
        self.load_file = None
        self.seed = seed
        self.initialize_seed()
        self.optimizer = optimizer

        # Define the matrices for the input images
        input_shape = (width, height, cells)
        left_input = Input(input_shape)
        right_input = Input(input_shape)

        # Get the CNN architecture as presented in the paper (read the readme for more information)
        model = self._get_architecture(input_shape)
        encoded_l = model(left_input)
        encoded_r = model(right_input)

        # Add a layer to combine the two CNNs
        L1_layer = Lambda(lambda tensors: K.abs(tensors[0] - tensors[1]))
        L1_siamese_dist = L1_layer([encoded_l, encoded_r])
        L1_siamese_dist = Dropout(dropout_rate)(L1_siamese_dist)

        # An output layer with Sigmoid activation function
        prediction = Dense(1, activation='sigmoid', bias_initializer=self.initialize_bias)(L1_siamese_dist)

        siamese_net = Model(inputs=[left_input, right_input], outputs=prediction)
        self.siamese_net = siamese_net
        self.siamese_net.compile(loss=loss, optimizer=optimizer, metrics=metrics)

    def initialize_seed(self):
        """
        Initialize seed all for environment
        """
        os.environ['PYTHONHASHSEED'] = str(self.seed)
        random.seed(self.seed)
        np.random.seed(self.seed)
        tf.random.set_seed(self.seed)

    def initialize_weights(self, shape, dtype=None):
        """
        Called when initializing the weights of the siamese model, uses the random_normal function of keras to return a
        tensor with a normal distribution of weights.
        """
        return K.random_normal(shape, mean=0.0, stddev=0.01, dtype=dtype, seed=self.seed)

    def initialize_bias(self, shape, dtype=None):
        """
        Called when initializing the biases of the siamese model, uses the random_normal function of keras to return a
        tensor with a normal distribution of weights.
        """
        return K.random_normal(shape, mean=0.5, stddev=0.01, dtype=dtype, seed=self.seed)

    def _get_architecture(self, input_shape):
        """
        Returns a Convolutional Neural Network based on the input shape given of the images. This is the CNN network
        that is used inside the siamese model. Uses parameters from the siamese one shot paper.
        """
        model = Sequential()
        model.add(
            Conv2D(filters=64,
                   kernel_size=(10, 10),
                   input_shape=input_shape,
                   kernel_initializer=self.initialize_weights,
                   kernel_regularizer=l2(2e-4),
                   name='Conv1'
                   ))
        model.add(BatchNormalization())
        model.add(Activation("relu"))
        model.add(MaxPooling2D())

        model.add(
            Conv2D(filters=128,
                   kernel_size=(7, 7),
                   kernel_initializer=self.initialize_weights,
                   bias_initializer=self.initialize_bias,
                   kernel_regularizer=l2(2e-4),
                   name='Conv2'
                   ))
        model.add(BatchNormalization())
        model.add(Activation("relu"))
        model.add(MaxPooling2D())

        model.add(
            Conv2D(filters=128,
                   kernel_size=(4, 4),
                   kernel_initializer=self.initialize_weights,
                   bias_initializer=self.initialize_bias,
                   kernel_regularizer=l2(2e-4),
                   name='Conv3'
                   ))
        model.add(BatchNormalization())
        model.add(Activation("relu"))
        model.add(MaxPooling2D())

        model.add(
            Conv2D(filters=256,
                   kernel_size=(4, 4),
                   kernel_initializer=self.initialize_weights,
                   bias_initializer=self.initialize_bias,
                   kernel_regularizer=l2(2e-4),
                   name='Conv4'
                   ))
        model.add(BatchNormalization())
        model.add(Activation("relu"))

        model.add(Flatten())
        model.add(
            Dense(4096,
                  activation='sigmoid',
                  kernel_initializer=self.initialize_weights,
                  kernel_regularizer=l2(2e-3),
                  bias_initializer=self.initialize_bias))
        return model

    def _load_weights(self, weights_file):
        """
        A function that attempts to load pre-existing weight files for the siamese model. If it succeeds then returns
        True and updates the weights, otherwise False.
        :return True if the file is already exists
        """
        # self.siamese_net.summary()
        self.load_file = weights_file
        if os.path.exists(weights_file):  # if the file is already exists, load and return true
            print('Loading pre-existed weights file')
            self.siamese_net.load_weights(weights_file)
            return True
        return False

    def fit(self, weights_file, train_path, validation_size, batch_size, epochs, early_stopping, patience, min_delta):
        """
        Function for fitting the model. If the weights already exist, just return the summary of the model. Otherwise,
        perform a whole train/validation/test split and train the model with the given parameters.
        """
        with open(train_path, 'rb') as f:
            x_train, y_train, names = pickle.load(f)
        """
        X_train[0]:  |----------x_train_0---------------------------|-------x_val_0--------|
        X_train[1]:  |----------x_train_1---------------------------|-------x_val_1--------|
        y_train:     |----------y_train_0 = y_train_1---------------|----y_val_0=y_val_1---|
        """
        x_train_0, x_val_0, y_train_0, y_val_0 = train_test_split(x_train[0], y_train,
                                                                  test_size=validation_size,
                                                                  random_state=self.seed)
        x_train_1, x_val_1, y_train_1, y_val_1 = train_test_split(x_train[1], y_train,
                                                                  test_size=validation_size,
                                                                  random_state=self.seed)
        x_train_0 = np.array(x_train_0, dtype='float64')
        x_val_0 = np.array(x_val_0, dtype='float64')
        x_train_1 = np.array(x_train_1, dtype='float64')
        x_val_1 = np.array(x_val_1, dtype='float64')
        x_train = [x_train_0, x_train_1]
        x_val = [x_val_0, x_val_1]
        if y_train_0 != y_train_1 and y_val_0 != y_val_1:
            raise Exception("y train lists or y validation list do not equal")
        y_train_both = np.array(y_train_0, dtype='float64')
        y_val_both = np.array(y_val_0, dtype='float64')
        if not self._load_weights(weights_file=weights_file):
            print('No such pre-existed weights file')
            print('Beginning to fit the model')
            callback = []
            if early_stopping:
                """
                We used the EarlyStopping function monitoring on the validation loss with a minimum delta of 0.1
                (Minimum change in the monitored quantity to qualify as an improvement, i.e.
                an absolute change of less than min_delta, will count as no improvement.) and patience 5 
                (Number of epochs with no improvement after which training will be stopped.).
                The direction is automatically inferred from the name of the monitored quantity (‘auto’).
                """
                es = EarlyStopping(monitor='val_loss', min_delta=min_delta, patience=patience, mode='auto', verbose=1)
                callback.append(es)
            self.siamese_net.fit(x_train, y_train_both, batch_size=batch_size, epochs=epochs,
                                 validation_data=(x_val, y_val_both), callbacks=callback, verbose=1)
            self.siamese_net.save_weights(self.load_file)
        # evaluate on the testing set
        loss, accuracy = self.siamese_net.evaluate(x_val, y_val_both, batch_size=batch_size)
        print(f'Loss on Validation set: {loss}')
        print(f'Accuracy on Validation set: {accuracy}')

    def evaluate(self, test_file, batch_size, analyze=False):
        """
        Function for evaluating the final model after training.
        test_file - file path to the test file.
        batch_size - the batch size used in training.

        Returns the loss and accuracy results.
        """
        with open(test_file, 'rb') as f:
            x_test, y_test, names = pickle.load(f)
        print(f'Available Metrics: {self.siamese_net.metrics_names}')
        y_test = np.array(y_test, dtype='float64')
        x_test[0] = np.array(x_test[0], dtype='float64')
        x_test[1] = np.array(x_test[1], dtype='float64')
        # evaluate on the test set
        loss, accuracy = self.siamese_net.evaluate(x_test, y_test, batch_size=batch_size)
        if analyze:
            self._analyze(x_test, y_test, names)
        return loss, accuracy

    def _analyze(self, x_test, y_test, names):
        """
        Function used for evaluating our network in the methods proposed in the assignment.
        We will find:
        - The person who has 2 images that are the most dissimilar to each other
        - The person with the two images that are the most similar to each other
        - Two people with the most dissimilar images, and
        - The two people with the most similar images.
        """
        best_class_0_prob = 1  # correct classification for different people, y=0, prediction->0
        best_class_0_name = None
        worst_class_0_prob = 0  # misclassification for different people, y=0, prediction->1
        worst_class_0_name = None
        best_class_1_prob = 0  # correct classification for same people, y=1, prediction->1
        best_class_1_name = None
        worst_class_1_prob = 1  # misclassification for same people, y=1, prediction->0
        worst_class_1_name = None
        prob = self.siamese_net.predict(x_test)
        for pair_index in range(len(names)):
            name = names[pair_index]
            y_pair = y_test[pair_index]
            pair_prob = prob[pair_index][0]
            if y_pair == 0:  # different people (actual)
                if pair_prob < best_class_0_prob:  # correct classification for different people, y=0, prediction->0
                    best_class_0_prob = pair_prob
                    best_class_0_name = name
                if pair_prob > worst_class_0_prob:  # misclassification for different people, y=0, prediction->1
                    worst_class_0_prob = pair_prob
                    worst_class_0_name = name
            else:  # the same person (actual)
                if pair_prob > best_class_1_prob:  # correct classification for same people, y=1, prediction->1
                    best_class_1_prob = pair_prob
                    best_class_1_name = name
                if pair_prob < worst_class_1_prob:  # misclassification for same people, y=1, prediction->0
                    worst_class_1_prob = pair_prob
                    worst_class_1_name = name

        print(f'correct classification for different people, y=0, prediction->0, name: {best_class_0_name} | prob: {best_class_0_prob}')
        print(f'misclassification for different people, y=0, prediction->1, name: {worst_class_0_name} | prob: {worst_class_0_prob}')
        print(f'correct classification for same people, y=1, prediction->1, name: {best_class_1_name} | prob: {best_class_1_prob}')
        print(f'misclassification for same people, y=1, prediction->0, name: {worst_class_1_name} | prob: {worst_class_1_prob}')


    def predict_stone_class(self, image, legend_patterns):
        """
        Function for predicting the stone class of an image.
        image - the image to predict the stone class of.

        Returns the predicted stone class.
        """
        # Prepraing the litho image for the network

        # Make de set of comparison with the legend images
        predictions = dict()
        for name,pattern in legend_patterns.items():
            prediction = self.siamese_net.predict([image,pattern])
            predictions[name] = prediction[0][0]
        # Keep the legend images with the highest probability
        return "CLAY"
    
print("Loaded Siamese Network")





class ResNetModel:
    """
        Create an ResNetModel object and methods to interact with it
    """

    model : None
    model_filepath : str
    data_dir : str = "./app/data/results/"
    # class_labels : list = ['anhydrite', 'calcareous_dolomite', 'chalk', 'chert', 'clay', 'coal_lignite', 'conglomerate', 'dolomite', 'dolomitic_limestone', 'fossiliferous', 'glauconite', 'gypsum', 'limestone', 'marl', 'metamorphic', 'pyrite', 'salt', 'sand', 'shale', 'silt', 'tuff']
    class_labels : list = ['coal_lignite', 'marl', 'marl_limestone', 'sand', 'sand_clay']
    
    def __init__(self, model_file_name : str):
        """
            Initialize the model
            args : model file path 
        """
        self.model_filepath = model_file_name
        self.model = tf.keras.models.load_model(self.model_filepath, compile=False)
        self.model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

    def predict(self, stones_directory : str):
        """
            Predict the lithology composition and measure the blocks heights
        """ 
        heights = {m: [] for m in self.class_labels}
        class_predictions = {m: 0 for m in self.class_labels}
        # Get the list of images
        image_files = glob.glob(stones_directory + "/*.png")
        for image_file in image_files:
            # Find image height
            img_original_size = Image.open(image_file)
            height = img_original_size.height
            valeur_relative = height / 100

            # Do the class prediction 
            img = load_img(image_file, target_size=(224, 224))
            x = img_to_array(img)
            x = np.expand_dims(x, axis=0)
            x = preprocess_input(x)
            preds = self.model.predict(x)
            class_idx = np.argmax(preds[0])
            predicted_label = self.class_labels[class_idx]
            print("Predicted : ", predicted_label)
            class_predictions[predicted_label] += 1
            heights[predicted_label].append(valeur_relative)
        return class_predictions, heights

    def create_excel(self, litho_dir : str, heights : dict, class_predictions : dict):
        """
            Create an excel file with the composition of the well
        """
        # Trouver la classe prédite qui a été choisie le plus grand nombre de fois
        most_predicted_class = max(class_predictions, key=class_predictions.get)
        
        # Trouver le nombre de fois que cette classe a été choisie
        most_predictions_count = class_predictions[most_predicted_class]

        # Créer le tableur Excel
        wb = openpyxl.Workbook()
        ws = wb.active

        # Ajouter les hauteurs relatives pour chaque matériau dans le tableur
        for i, m in enumerate(self.class_labels):
            # Ajouter le nom du matériau dans la première colonne
            ws.cell(row=i+2, column=1, value=m)

        # Ajouter les hauteurs relatives dans les colonnes suivantes
        for j, h in enumerate(heights[m]):
            ws.cell(row=i+2, column=j+2, value=h)

        # Ajouter la ligne des totaux en bas
        #for j in range(2, 7):
        #10 etant le nb de fois qu'un matériau est prédit
        for j in range(2, most_predictions_count+2):
            sub_sum = sum(ws.cell(row=i, column=j).value or 0 for i in range(2, len(self.class_labels)+2))
            ws.cell(row=len(self.class_labels)+2, column=j, value=sub_sum)

        # Ajouter la colonne des totaux
        for i in range(2, len(self.class_labels)+3):
            sub_sum = sum(ws.cell(row=i, column=j).value or 0 for j in range(2,  len(self.class_labels)))
            ws.cell(row=i, column=most_predictions_count+2, value=sub_sum)

        # Ajouter la colonne des pourcentages
        for i in range(2, len(self.class_labels)+2):
            total = ws.cell(row=len(self.class_labels)+2, column=most_predictions_count+2).value
            percent = sum(ws.cell(row=i, column=j).value or 0 for j in range(2, most_predictions_count+2)) / ws.cell(row=len(self.class_labels)+2, column=most_predictions_count+2).value
            ws.cell(row=i, column=most_predictions_count+3, value=percent).number_format = '0.00%'

        #génération du graphique camembert
        # Créer le graphique camembert
        chart = PieChart()
        chart.title = "Pourcentage Lithologie"

        # Ajouter les données du graphique
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(self.class_labels)+1)
        data = Reference(ws, min_col=most_predictions_count+3, min_row=1, max_row=len(self.class_labels)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        # Ajouter le graphique au tableur
        ws.add_chart(chart, "J2")

        # Enregistrer le tableur Excel
        wb.save(litho_dir+"resultats.xlsx")

        return 1

    def get_composition_from_excel(self,litho_dir : str):
        composition_dict = {}
        res_path = litho_dir+"resultats.xlsx"
        # Nom de la feuille dans le fichier Excel
        nom_feuille = "Sheet"

        # Lire le fichier Excel en utilisant pandas
        df = pd.read_excel(res_path, sheet_name=nom_feuille)

        # Cette colonne sera toujours en premièr, et donc s'appellera toujours comme ca
        # Nom de la colonne que vous voulez retourner
        nom_colonne = "Unnamed: 0"

        # Sélectionner la colonne spécifique en utilisant le nom de la colonne
        df[nom_colonne] = df[nom_colonne].dropna()
        list_nom = df[nom_colonne].to_list()
        # Supprime le denrier élément de la liste
        # Comme on ajoute une ligne à la fin du tableau pour les totaux, cela sera toujours de cette configuration aussi
        list_nom.pop()
        # print(df)

        # Afficher la colonne
        # print(list_nom)

        # Supprimer toutes les colonnes qui contiennent plus de 5 NaN
        df = df.dropna(thresh=len(df)-(len(list_nom) - 1) , axis = 1)

        # Sélectionner la dernière colonne restante
        list_nombre = df.iloc[:, -2].to_list()
        list_nombre.pop()
        # print(list_nombre)
        return list_nom, list_nombre
    
    def composition(self,well_name : str):
        """
            Get the composition of the well based on a file

            Args: file is a File object

            Returns: the dictionnary of the compositions
        """ 
        litho = Lithologie()
        litho.split_litho(well_name,"./app/data/results/"+well_name+"/","./app/data/results/"+well_name+"/stones")
        stones_directory = litho.litho_stones_dir
        class_predictions, heights = self.predict(stones_directory)
        self.create_excel(litho.litho_dir,heights,class_predictions)
        noms, valeurs = self.get_composition_from_excel(litho.litho_dir)
        return noms, valeurs
    



   